from numpy import less
import openpyxl
import random
import pandas as pd
import csv
import math

Notesarray = [[2000, 500, 100], [2000, 500, 200, 100]]
Notes = dict()
arr = random.choice(Notesarray)


for Note in arr:
    n = random.randint(500, 1000)
    Notes.update({Note: n})


details = dict()


book = openpyxl.load_workbook(
    "C:\\Users\\preetham\\Documents\\CL intern projects\\ATM Project\\ATM_python_file.xlsx"
)
sheet = book.active


def numberOfNotes(amount):
    answer = dict()
    if len(Notes) == 3:
        answer.update({2000: math.floor(amount / 2000)})
        amount = amount % 2000
        answer.update({500: math.floor(amount / 500)})
        amount = amount % 500
        answer.update({100: math.floor(amount / 100)})
        amount = amount % 100

    else:
        answer.update({2000: math.floor(amount / 2000)})
        amount = amount % 2000
        answer.update({500: math.floor(amount / 500)})
        amount = amount % 500
        answer.update({200: math.floor(amount / 200)})
        amount = amount % 200
        answer.update({100: math.floor(amount / 100)})
        amount = amount % 100
    return answer


def Withdraw():
    amount = int(input("Enter amount to withdraw\n"))
    if amount % 100 > 0:
        print("Invalid entry:Enter atleast multiple of 100\n")
        return

    if amount > details["Balance"]:
        print("Insufficient balance")
    else:
        pin = int(input("Enter pin to confirm\n"))

        if pin == details["PIN"]:

            notesDrawn = numberOfNotes(amount)
            details["Balance"] = details["Balance"] - amount
            sheet.cell(row=2, column=4).value = details["Balance"]

            for notes in notesDrawn:
                print(str(notes) + " x " + str(notesDrawn[notes]) + "\n")
            print("Cash withdrawn!")
        else:
            print("Wrong pin")

    return


def Deposite():
    amount = int(input("Enter amount to Deposit\n"))
    details["Balance"] = details["Balance"] + amount
    sheet.cell(row=2, column=4).value = details["Balance"]
    return


def Check_Balance():
    print(details["Balance"])
    return


def Change_PIN():
    pin = int(input("Enter new pin\n"))
    while pin < 1000 or pin > 9999:
        pin = int(input("Wrong input! Enter 4 Digits only"))

    cpin = int(input("confirm pin\n"))
    if pin == cpin:
        details["PIN"] = pin
        sheet.cell(row=2, column=3).value = details["PIN"]
        print("Confirmation successfull")
    else:
        print("Confirmation failed!\n")
    return


def checkPIn(PIN):
    if PIN < 1000 or PIN > 9999:
        pin = "Inavlid input: Enter 4 Digits\n"
        checkPIn(pin)
        return
    if PIN == details["PIN"]:
        print("Succesfull login!")
        while True:
            option = int(
                input(
                    "Choose option 1.Withdraw 2.Deposite 3.Check Balance 4.Help 5.Done\n"
                )
            )
            if option == 1:
                Withdraw()
            elif option == 2:
                Deposite()
            elif option == 3:
                Check_Balance()
            elif option == 4:
                Change_PIN()
            elif option == 5:
                print("Thank you!")
                break
            else:
                print("Invalid entry,Please try again\n")
    else:
        pin = int(input("Unauthorized User,Enter correct PIN.\n"))
        checkPIn(pin)
        return


def main():

    row1 = sheet[1]
    row2 = sheet[2]
    size = len(row1)

    for i in range(size):
        details.update({row1[i].value: row2[i].value})

    pin = int(input("Enter your PIN\n"))
    checkPIn(pin)

    book.save(
        "C:\\Users\\preetham\\Documents\\CL intern projects\\ATM Project\\ATM_python_file.xlsx"
    )

    # f = open(
    #     "C:\\Users\\preetham\\Documents\\CL intern projects\\ATM Project\\ATM_python_file.csv",
    #     "r",
    # )

    # reader = csv.reader(f)
    # data = pd.read_csv(
    #     "C:\\Users\\preetham\\Documents\\CL intern projects\\ATM Project\\ATM_python_file.csv"
    # )

    # header = list()
    # for info in data:
    #     header.append(info)

    # for i in range(len(header)):
    #     details.update({header[i]: data[header[i]].tolist()[0]})

    # df = pd.DataFrame(details, index=[0])
    # df.to_csv(
    #     "C:\\Users\\preetham\\Documents\\CL intern projects\\ATM Project\\ATM_python_file4.csv"
    # )

    # for d in df:
    #     print(d)

    # for rows in reader:
    #     row1 = rows

    #     print(row1)

    #     details.update({row1[i].value: row2[i].value})
    # print(details)


# print(dict_from_csv)
#     print(type(df))
#     print(type(df.head()))
#     for d in df.head():
#         print(d)
# pin = int(input("Enter pin\n"))
# checkPIn(pin)


if __name__ == "__main__":
    main()

    # print(row1)

    #         details.update({cell: row[1]})
    # details.update({cell.value: ""})
    # print(details)

    # if details.get("PIN ") == "No PIN":
    #     pin = input("Create a PIN please")
    #     cpin= input("confirm ")


# f1 = open("C:\\Users\\preetham\\Documents\\ATM python file.xlsx")
# workbook = xlsxwriter.Workbook(f1)
# worksheet = workbook.add_worksheet()

# worksheet.write(0, 0, 1234)  # Writes an int
# worksheet.write(1, 0, 1234.56)  # Writes a float
# worksheet.write(2, 0, "Hello")  # Writes a string
# worksheet.write(3, 0, None)  # Writes None
# worksheet.write(4, 0, True)


# f = open(ATM_python_file.xlsx)


# csvreader = csv.reader(f)

# details = dict()

# for row in csvreader:
#     details.update({row[0]: row[1]})

# print(details)

# if details.get("PIN ") == "No PIN":
#     pin = input("Provide PIN Please")
#     # print(pin)
#     csvwriter = csv.writer(f)
#     csvwriter.writerow(["PIn", pin])


# if rows[1][1] == "No PIN":
#     pin = input("Set up a PIN")
#     df = pd.read_csv("C:\\Users\\preetham\\Documents\\ATM python file.csv")
#     df.loc[1, "Preetham Wilson Saldanha"] = pin
#     df.to_csv("C:\\Users\\preetham\\Documents\\ATM python file.csv")
#     print(df)


# book.save(
#     "C:\\Users\\preetham\\Documents\\CL intern projects\\ATM Project\\ATM_python_file.xlsx"
# )
